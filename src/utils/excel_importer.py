import csv
import os
from typing import Dict, List, Any

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def parse_account_file(file_path: str) -> Dict[str, Any]:
    """
    Parses an Excel (.xlsx) or CSV (.csv) file for account and character data.
    
    Expected Format:
    Row 1: Email (in first cell)
    Row 2: Header (ignored)
    Rows 3+: Cantidad (Slots), Nombre PJ
    
    Returns:
        Dict: {
            "email": "...",
            "characters": [{"slots": int, "name": str}, ...]
        }
    """
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.csv':
        return _parse_csv(file_path)
    elif ext == '.xlsx':
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required to parse .xlsx files. Please install it or use .csv")
        return _parse_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

def _parse_csv(file_path: str) -> Dict[str, Any]:
    data = {"email": "", "characters": []}
    
    with open(file_path, mode='r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
        
        if not rows:
            return data
            
        # Email is in first row
        data["email"] = rows[0][0].strip()
        
        # Characters start from row 3 (index 2)
        # Skip header in row 2
        for row in rows[2:]:
            if len(row) >= 2:
                try:
                    slots = int(row[0])
                    name = row[1].strip()
                    if name:
                        data["characters"].append({"slots": slots, "name": name})
                except ValueError:
                    continue # Skip invalid rows
                    
    return data

def _parse_xlsx(file_path: str) -> Dict[str, Any]:
    data = {"email": "", "characters": []}
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active # Use the first sheet
    
    # Email in cell A1
    data["email"] = str(ws.cell(row=1, column=1).value or "").strip()
    
    # Characters start from row 3
    # Column A: Cantidad (Slots), Column B: Nombre PJ
    for row_idx in range(3, ws.max_row + 1):
        slots_val = ws.cell(row=row_idx, column=1).value
        name_val = ws.cell(row=row_idx, column=2).value
        
        if name_val:
            try:
                slots = int(slots_val) if slots_val is not None else 5 # Default to 5
                name = str(name_val).strip()
                data["characters"].append({"slots": slots, "name": name})
            except (ValueError, TypeError):
                continue
                
    return data
