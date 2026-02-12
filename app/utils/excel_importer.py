import csv
import os
from typing import Dict, List, Any
from app.utils.logger import logger

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def parse_account_file(file_path: str) -> Dict[str, Any]:
    """
    Parses an Excel (.xlsx) or CSV (.csv) file for account and character data.
    
    Expected Format:
    Row 1: Email (in first cell)
    Row 2: Header (ignored)
    Rows 3+: Cantidad (Slots), Nombre PJ
    
    Returns:
        Dict: {
            "email": "...",
            "characters": [{"slots": int, "name": str}, ...]
        }
    """
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.csv':
        return _parse_csv(file_path)
    elif ext == '.xlsx':
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required to parse .xlsx files. Please install it or use .csv")
        return _parse_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

def _parse_csv(file_path: str) -> Dict[str, Any]:
    data = {"email": "", "characters": []}
    
    with open(file_path, mode='r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
        
        if not rows:
            return data
            
        # Email is in first row
        raw_email = rows[0][0].strip()
        if raw_email and "@" not in raw_email:
            data["email"] = f"{raw_email}@gmail.com"
        else:
            data["email"] = raw_email
        
        # Characters start from row 3 (index 2)
        # Skip header in row 2
        for row in rows[2:]:
            if len(row) >= 3:
                try:
                    # Row: [Account, Slots, Name]
                    account_name = row[0].strip()
                    slots = int(row[1])
                    name = row[2].strip()
                    if name and account_name:
                        data["characters"].append({
                            "slots": slots, 
                            "name": name,
                            "account_name": account_name
                        })
                except ValueError:
                    continue # Skip invalid rows
                    
    return data

def _parse_xlsx(file_path: str) -> Dict[str, Any]:
    data = {"email": "", "characters": []}
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active # Use the first sheet
    
    # Smart Parser
    # 1. Identify Header Row
    header_row_idx = -1
    slots_col_idx = -1
    name_col_idx = -1
    
    # We load all rows into a list of list of strings to search easier
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c).strip() if c is not None else "" for c in row])
        
    for i, row_strs in enumerate(all_rows):
        if i > 10: break
        row_lower = [s.lower() for s in row_strs]
        
        # Format A: Standard (User Screenshot)
        # Headers: Cantidad (Slots), Alquimero/Pj (Name)
        if "cantidad" in row_lower:
            header_row_idx = i
            for idx, val in enumerate(row_lower):
                if "cantidad" in val: slots_col_idx = idx
                elif any(x in val for x in ["alquimero", "pj", "personaje", "mezclador", "nombre"]): name_col_idx = idx
            break
            
        # Format B: Legacy/Weird (Cors Diarios.xlsx)
        # Headers: Pj (Slots), Fragmentos (Name), Cuenta (Account)
        elif "fragmentos" in row_lower and "pj" in row_lower:
            header_row_idx = i
            for idx, val in enumerate(row_lower):
                if "pj" in val: slots_col_idx = idx     # In this file, Pj is Slots (Col B)
                elif "fragmentos" in val: name_col_idx = idx # Fragmentos is Name (Col C)
            break
            
    # 2. Identify Email (Store Name)
    # Scan rows 0-5. Ignore Header Row. Ignore Rows with Dates.
    found_email = ""
    start_data_row = 3 # Default legacy
    
    if header_row_idx != -1:
        # Check rows around header
        # Candidates: Rows 0..header_row_idx..header_row_idx+2
        # Just scan first 10 rows basically, but skip header
        for i in range(min(10, len(all_rows))):
             if i == header_row_idx: continue
             
             # Check logic: Skip Dates, look for single string
             row_vals = [x for x in all_rows[i] if x]
             if not row_vals: continue
             
             txt = row_vals[0] # Take first non-empty
             
             # Skip Dates
             if "/" in txt or "-" in txt or "202" in txt: # 2025, 2024
                 continue
                 
             # Skip if it looks like header keywords
             if "cantidad" in txt.lower() or "pj" in txt.lower(): continue
             
             # Skip if it looks like Data (e.g. "Fragmetin1", "5", "Name")
             # If row has multiple values, it might be data?
             # User Email "FRAGMETIN5" is usually one Merged cell (1 value) OR repeated?
             
             if len(txt) > 2:
                 found_email = txt
                 break
                 
        start_data_row = header_row_idx + 2 # Start 1 row after header (1-based)
        
    if found_email:
         if "@" not in found_email:
            data["email"] = f"{found_email}@gmail.com"
         else:
            data["email"] = found_email
    else:
         # Legacy fallback
         data["email"] = str(ws.cell(row=1, column=1).value or "").strip()
         if data["email"] and "@" not in data["email"] and len(data["email"]) > 2 and "/" not in data["email"] and "-" not in data["email"]:
              data["email"] += "@gmail.com"

    # 3. Parse Data
    # Use detected columns or defaults (Col 2/3 mapped from 1/2) 
    # ws.cell is 1-based. 
    # Account is ALWAYS Column A (1).
    if slots_col_idx == -1: slots_col_idx = 1 # Col B (0-based index 1)
    if name_col_idx == -1: name_col_idx = 2   # Col C (0-based index 2)
    
    for row_idx in range(start_data_row, ws.max_row + 1):
        # ws.cell indices
        try:
             # Check if row is valid data (has values in specific account column?)
             account_val = ws.cell(row=row_idx, column=1).value
             
             # Parsing Logic ...
             name_val = ws.cell(row=row_idx, column=name_col_idx + 1).value
             
             if not name_val: continue # Skip if no name
             
             slots_val = ws.cell(row=row_idx, column=slots_col_idx + 1).value
             
             slots = int(slots_val) if slots_val is not None else 5
             name = str(name_val).strip()
             account_name = str(account_val).strip() if account_val else ""
             
             if name:
                 data["characters"].append({
                        "slots": slots, 
                        "name": name,
                        "account_name": account_name
                 })
        except:
            continue
            
    return data
                
    return data
