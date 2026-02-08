import openpyxl
from app.infrastructure.database.setup import init_db
from app.domain.models import StoreAccount, GameAccount, Character, CharacterType
from app.utils.config import Config
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

class ImportService:
    def __init__(self, session=None):
        if session:
            self.session = session
        else:
            engine = create_engine(Config.get_db_url())
            Session = sessionmaker(bind=engine)
            self.session = Session()

    def import_from_excel(self, file_path, server_id):
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            processed_count = 0
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                current_store_email = None
                current_store_account = None
                
                # Scan rows
                # Using iter_rows to get cell objects directly for style check
                for row in ws.iter_rows(min_row=1):
                    if not row:
                        continue
                        
                    first_cell = row[0]
                    if not first_cell:
                        continue
                        
                    # Check Background Color for Header (Yellow)
                    # "FFFF00" is usually standard yellow. Excel might use AARRGGBB.
                    # We check if "FFFF00" matches vaguely or specific codes.
                    is_yellow = False
                    if first_cell.fill and first_cell.fill.fgColor and first_cell.fill.fgColor.rgb:
                        color_hex = str(first_cell.fill.fgColor.rgb) # Could be 'FFFFFF00'
                        if "FFFF00" in color_hex or color_hex == "FFFF0000": # Handle variations
                            is_yellow = True
                    
                    val0 = first_cell.value
                    
                    if is_yellow and val0:
                        # New Store Account Section
                        # Clean value just in case
                        base_name = str(val0).strip()
                        # If user didn't put @gmail.com, we append it as per requirement
                        if "@" not in base_name:
                            email = f"{base_name}@gmail.com"
                        else:
                            email = base_name
                        
                        current_store_email = email
                        
                        # Find or Create StoreAccount
                        store_acc = self.session.query(StoreAccount).filter_by(email=email).first()
                        if not store_acc:
                            store_acc = StoreAccount(email=email)
                            self.session.add(store_acc)
                            self.session.flush() # Get ID
                        
                        current_store_account = store_acc
                        
                    elif current_store_account and val0:
                        # Data Row: [GameAccName, Count, AlchemistName, ...]
                        game_acc_name = str(val0).strip()
                        
                        # Column 1: Count (Slots)
                        count_val = None
                        if len(row) > 1 and row[1].value:
                            count_val = row[1].value
                        
                        try:
                            total_slots = int(count_val) if count_val else 5
                        except:
                            total_slots = 5

                        # Column 2: Alchemist Name
                        alchemist_name = None
                        if len(row) > 2 and row[2].value:
                            alchemist_name = str(row[2].value).strip()
                            
                        if not game_acc_name or not alchemist_name:
                            continue
                            
                        # Find or Create GameAccount
                        # Check global uniqueness of username
                        game_acc = self.session.query(GameAccount).filter_by(username=game_acc_name).first()
                        
                        if game_acc:
                            # Account exists: Update links if necessary
                            if game_acc.store_account_id != current_store_account.id:
                                game_acc.store_account_id = current_store_account.id
                            if game_acc.server_id != server_id:
                                game_acc.server_id = server_id
                        else:
                            # Create new
                            game_acc = GameAccount(
                                username=game_acc_name,
                                server_id=server_id,
                                store_account_id=current_store_account.id
                            )
                            self.session.add(game_acc)
                            self.session.flush()
                        
                        # 1. Ensure Alchemist Character exists
                        alchemist_exists = False
                        for char in game_acc.characters:
                            if char.name == alchemist_name:
                                alchemist_exists = True
                                break
                        
                        if not alchemist_exists:
                            new_char = Character(
                                name=alchemist_name,
                                char_type=CharacterType.ALCHEMIST,
                                game_account_id=game_acc.id
                            )
                            self.session.add(new_char)
                            processed_count += 1
                        
                        # 2. Fill with Dummy Characters up to 'total_slots'
                        # We use the Alchemist + Dummies to reach total
                        # Refresh characters list
                        self.session.refresh(game_acc) # Ensure we have latest list
                        current_count = len(game_acc.characters)
                        
                        if current_count < total_slots:
                            needed = total_slots - current_count
                            for i in range(needed):
                                # Generate a generic name "PJ_X_<username>"
                                # Use high index or just safe random
                                # We can just use "PJ {i+2}" logic or similar
                                
                                # Simple safe generic name:
                                idx = current_count + i + 1
                                dummy_name = f"PJ_{idx}_{game_acc.username}"
                                
                                dummy = Character(
                                    name=dummy_name,
                                    char_type=CharacterType.ALCHEMIST, # Or other? Default Alchemist for now
                                    game_account_id=game_acc.id
                                )
                                self.session.add(dummy)

            self.session.commit()
            return {"success": True, "processed_accounts": processed_count}

        except Exception as e:
            self.session.rollback()
            return {"success": False, "error": str(e)}
        finally:
            # If we created the session, we should close it. 
            # If passed in, caller handles it. But here we assume we manage it if we created it?
            # For simplicity in this controller, we can leave it open if passed, 
            # but usually good practice to close if we opened it.
            pass 
