import openpyxl
from openpyxl.utils import get_column_letter

file_path = "Cors Diarios (2).xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Loaded {file_path}")
    print(f"Sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        print(f"\nScanning Sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        for row in ws.iter_rows(min_row=1, max_row=50): # Scan first 50 rows
            row_idx = row[0].row
            first_cell = row[0]
            
            # Check background color
            fill = first_cell.fill
            # Typical Yellow is FFFF0000 (ARGB) or FFFF00
            bg_color = fill.fgColor.rgb if fill.fgColor else "None"
            
            values = [c.value for c in row[:5]] # Print first 5 columns
            
            if bg_color and "FFFF00" in str(bg_color):
                 print(f"YELLOW ROW {row_idx}: Color={bg_color} Values={values}")
            elif values[0] is not None:
                 print(f"Data Row {row_idx}: Values={values}")

except Exception as e:
    print(f"Error: {e}")
